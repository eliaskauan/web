# -*- coding: utf-8 -*-
"""
Processador de arquivos Excel para o sistema de web scraping Parts Unlimited
"""

import logging
import os
from typing import List, Tuple, Dict, Optional
from pathlib import Path
import time

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelProcessor:
    """
    Classe responsável pelo processamento de arquivos Excel
    """
    
    def __init__(self, excel_path: str):
        """
        Inicializa o processador Excel
        
        Args:
            excel_path: Caminho para o arquivo Excel
        """
        self.excel_path = Path(excel_path)
        self.workbook: Optional[Workbook] = None
        self.sheet_produtos = None
        self.sheet_credenciais = None
        self.sheet_configuracoes = None
        self.logger = logging.getLogger(__name__)
        
        # Configurações padrão das colunas
        self.colunas_produtos = {
            'A': 'codigo_produto',
            'B': 'link_produto',
            'C': 'descricao_titulo',
            'D': 'sub_descricao',
            'E': 'features',
            'F': 'specs',
            'G': 'part_codes',
            'H': 'part_notices',
            'I': 'certifications',
            'J': 'references',
            'K': 'package_info',
            'L': 'size_chart',
            'M': 'video_url',
            'N': 'imagens_urls',
            'O': 'substituicao_oem',
            'P': 'tabela_ajustes',
            'Q': 'texto_ajustes',
            'R': 'link_catalogo',
            'S': 'imagem_diretorio',
            'T': 'video_detalhado',
            'U': 'status_processamento',
            'V': 'data_processamento',
            'W': 'observacoes'
        }
        
        # Verificar se arquivo existe, se não existir, criar um novo
        if not self.excel_path.exists():
            self.criar_planilha_padrao()
        
        self.carregar_excel()
    
    def criar_planilha_padrao(self):
        """Cria uma planilha Excel padrão com as abas necessárias"""
        try:
            # Criar workbook
            wb = Workbook()
            
            # Remover aba padrão
            wb.remove(wb.active)
            
            # Criar aba de produtos
            ws_produtos = wb.create_sheet("Produtos")
            self._criar_cabecalho_produtos(ws_produtos)
            
            # Criar aba de credenciais
            ws_credenciais = wb.create_sheet("Credenciais")
            self._criar_aba_credenciais(ws_credenciais)
            
            # Criar aba de configurações
            ws_config = wb.create_sheet("Configuracoes")
            self._criar_aba_configuracoes(ws_config)
            
            # Salvar arquivo
            wb.save(self.excel_path)
            self.logger.info(f"Planilha padrão criada: {self.excel_path}")
            
        except Exception as e:
            self.logger.error(f"Erro ao criar planilha padrão: {e}")
            raise
    
    def _criar_cabecalho_produtos(self, worksheet):
        """Cria o cabeçalho da aba de produtos"""
        cabecalhos = [
            'Código do Produto',
            'Link do Produto',
            'Descrição Título',
            'Sub Descrição',
            'Features',
            'Specs',
            'Part Codes',
            'Part Notices',
            'Certifications',
            'References',
            'Package Info',
            'Size Chart',
            'Vídeo URL',
            'Imagens URLs',
            'Substituição OEM',
            'Tabela Ajustes',
            'Texto Ajustes',
            'Link Catálogo',
            'Imagem Diretório',
            'Vídeo Detalhado',
            'Status',
            'Data Processamento',
            'Observações'
        ]
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, cabecalho in enumerate(cabecalhos, 1):
            cell = worksheet.cell(row=1, column=col, value=cabecalho)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
            # Ajustar largura da coluna
            worksheet.column_dimensions[get_column_letter(col)].width = 20
    
    def _criar_aba_credenciais(self, worksheet):
        """Cria a aba de credenciais"""
        worksheet['A1'] = 'Configuração de Login'
        worksheet['A1'].font = Font(bold=True, size=14)
        
        worksheet['A3'] = 'URL Login:'
        worksheet['B3'] = 'https://www.parts-unlimited.com/login'
        
        worksheet['A4'] = 'Username:'
        worksheet['B4'] = ''  # Usuário preencherá
        
        worksheet['A5'] = 'Password:'
        worksheet['B5'] = ''  # Usuário preencherá
        
        worksheet['A7'] = 'Seletores CSS:'
        worksheet['A8'] = 'Campo Username:'
        worksheet['B8'] = '#username, input[name="username"], input[type="email"]'
        
        worksheet['A9'] = 'Campo Password:'
        worksheet['B9'] = '#password, input[name="password"], input[type="password"]'
        
        worksheet['A10'] = 'Botão Login:'
        worksheet['B10'] = 'button[type="submit"], input[type="submit"], .login-btn'
        
        # Ajustar larguras
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 50
    
    def _criar_aba_configuracoes(self, worksheet):
        """Cria a aba de configurações"""
        worksheet['A1'] = 'Configurações do Scraper'
        worksheet['A1'].font = Font(bold=True, size=14)
        
        configuracoes = [
            ('Diretório Imagens:', 'images/'),
            ('Diretório Vídeos:', 'videos/'),
            ('Delay Mínimo (seg):', '2'),
            ('Delay Máximo (seg):', '8'),
            ('Max Tentativas:', '3'),
            ('Timeout (ms):', '30000'),
            ('Salvar a cada N produtos:', '5'),
            ('Criar backup:', 'SIM'),
            ('Modo Headless:', 'SIM'),
            ('Debug:', 'NAO')
        ]
        
        for i, (config, valor) in enumerate(configuracoes, 3):
            worksheet[f'A{i}'] = config
            worksheet[f'B{i}'] = valor
        
        # Ajustar larguras
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 20
    
    def carregar_excel(self):
        """Carrega o arquivo Excel"""
        try:
            self.workbook = load_workbook(self.excel_path)
            
            # Carregar abas
            if "Produtos" in self.workbook.sheetnames:
                self.sheet_produtos = self.workbook["Produtos"]
            else:
                raise ValueError("Aba 'Produtos' não encontrada")
            
            if "Credenciais" in self.workbook.sheetnames:
                self.sheet_credenciais = self.workbook["Credenciais"]
            
            if "Configuracoes" in self.workbook.sheetnames:
                self.sheet_configuracoes = self.workbook["Configuracoes"]
            
            self.logger.info(f"Excel carregado: {self.excel_path}")
            
        except Exception as e:
            self.logger.error(f"Erro ao carregar Excel: {e}")
            raise
    
    def obter_credenciais(self) -> Dict[str, str]:
        """
        Obtém as credenciais de login da aba credenciais
        
        Returns:
            Dicionário com credenciais
        """
        credenciais = {}
        
        try:
            if self.sheet_credenciais:
                credenciais = {
                    'url_login': self.sheet_credenciais['B3'].value or '',
                    'username': self.sheet_credenciais['B4'].value or '',
                    'password': self.sheet_credenciais['B5'].value or '',
                    'selector_username': self.sheet_credenciais['B8'].value or '',
                    'selector_password': self.sheet_credenciais['B9'].value or '',
                    'selector_login_btn': self.sheet_credenciais['B10'].value or ''
                }
            
            return credenciais
            
        except Exception as e:
            self.logger.error(f"Erro ao obter credenciais: {e}")
            return {}
    
    def obter_configuracoes(self) -> Dict[str, str]:
        """
        Obtém as configurações da aba configurações
        
        Returns:
            Dicionário com configurações
        """
        config = {}
        
        try:
            if self.sheet_configuracoes:
                # Mapear células para configurações
                config = {
                    'diretorio_imagens': self.sheet_configuracoes['B3'].value or 'images/',
                    'diretorio_videos': self.sheet_configuracoes['B4'].value or 'videos/',
                    'delay_min': int(self.sheet_configuracoes['B5'].value or 2),
                    'delay_max': int(self.sheet_configuracoes['B6'].value or 8),
                    'max_tentativas': int(self.sheet_configuracoes['B7'].value or 3),
                    'timeout': int(self.sheet_configuracoes['B8'].value or 30000),
                    'salvar_cada_n': int(self.sheet_configuracoes['B9'].value or 5),
                    'criar_backup': str(self.sheet_configuracoes['B10'].value or 'SIM').upper() == 'SIM',
                    'headless': str(self.sheet_configuracoes['B11'].value or 'SIM').upper() == 'SIM',
                    'debug': str(self.sheet_configuracoes['B12'].value or 'NAO').upper() == 'SIM'
                }
            
            return config
            
        except Exception as e:
            self.logger.error(f"Erro ao obter configurações: {e}")
            return {}
    
    def obter_produtos_pendentes(self) -> List[Tuple[int, str]]:
        """
        Obtém lista de produtos que ainda não foram processados
        
        Returns:
            Lista de tuplas (linha, codigo_produto)
        """
        produtos_pendentes = []
        
        try:
            if not self.sheet_produtos:
                return produtos_pendentes
            
            # Iterar pelas linhas da planilha (começando da linha 2, pois linha 1 é cabeçalho)
            for row in range(2, self.sheet_produtos.max_row + 1):
                codigo_produto = self.sheet_produtos[f'A{row}'].value
                status = self.sheet_produtos[f'U{row}'].value
                
                # Se tem código do produto e não foi processado
                if codigo_produto and (not status or status.strip() == ''):
                    produtos_pendentes.append((row, str(codigo_produto).strip()))
            
            self.logger.info(f"Encontrados {len(produtos_pendentes)} produtos pendentes")
            return produtos_pendentes
            
        except Exception as e:
            self.logger.error(f"Erro ao obter produtos pendentes: {e}")
            return []
    
    def atualizar_produto(self, linha: int, dados_produto: Dict[str, str]):
        """
        Atualiza os dados de um produto na planilha
        
        Args:
            linha: Número da linha na planilha
            dados_produto: Dicionário com dados do produto
        """
        try:
            if not self.sheet_produtos:
                raise ValueError("Aba de produtos não carregada")
            
            # Mapear dados para colunas
            mapeamento = {
                'link_produto': 'B',
                'descricao_titulo': 'C',
                'sub_descricao': 'D',
                'features': 'E',
                'specs': 'F',
                'part_codes': 'G',
                'part_notices': 'H',
                'certifications': 'I',
                'references': 'J',
                'package_info': 'K',
                'size_chart': 'L',
                'video_url': 'M',
                'imagens_urls': 'N',
                'substituicao_oem': 'O',
                'tabela_ajustes': 'P',
                'texto_ajustes': 'Q',
                'link_catalogo': 'R',
                'imagem_diretorio': 'S',
                'video_detalhado': 'T'
            }
            
            # Atualizar dados do produto
            for campo, coluna in mapeamento.items():
                if campo in dados_produto:
                    valor = dados_produto[campo]
                    # Se for lista, converter para string separada por ;
                    if isinstance(valor, list):
                        valor = '; '.join(str(v) for v in valor)
                    
                    self.sheet_produtos[f'{coluna}{linha}'].value = str(valor) if valor else ''
            
            # Atualizar status e data
            self.sheet_produtos[f'U{linha}'].value = 'CONCLUIDO'
            self.sheet_produtos[f'V{linha}'].value = time.strftime('%Y-%m-%d %H:%M:%S')
            
            self.logger.debug(f"Produto atualizado na linha {linha}")
            
        except Exception as e:
            self.logger.error(f"Erro ao atualizar produto na linha {linha}: {e}")
            raise
    
    def marcar_produto_nao_encontrado(self, linha: int):
        """
        Marca um produto como não encontrado
        
        Args:
            linha: Número da linha na planilha
        """
        try:
            if self.sheet_produtos:
                self.sheet_produtos[f'U{linha}'].value = 'NAO_ENCONTRADO'
                self.sheet_produtos[f'V{linha}'].value = time.strftime('%Y-%m-%d %H:%M:%S')
                self.sheet_produtos[f'W{linha}'].value = 'Produto não encontrado no site'
                
        except Exception as e:
            self.logger.error(f"Erro ao marcar produto como não encontrado: {e}")
    
    def marcar_produto_erro(self, linha: int, erro: str):
        """
        Marca um produto com erro
        
        Args:
            linha: Número da linha na planilha
            erro: Descrição do erro
        """
        try:
            if self.sheet_produtos:
                self.sheet_produtos[f'U{linha}'].value = 'ERRO'
                self.sheet_produtos[f'V{linha}'].value = time.strftime('%Y-%m-%d %H:%M:%S')
                self.sheet_produtos[f'W{linha}'].value = str(erro)[:500]  # Limitar tamanho do erro
                
        except Exception as e:
            self.logger.error(f"Erro ao marcar produto com erro: {e}")
    
    def salvar_excel(self, backup: bool = True):
        """
        Salva o arquivo Excel
        
        Args:
            backup: Se True, cria backup antes de salvar
        """
        try:
            if not self.workbook:
                raise ValueError("Workbook não carregado")
            
            # Criar backup se solicitado
            if backup and self.excel_path.exists():
                backup_path = self.excel_path.with_suffix(f'.backup_{int(time.time())}.xlsx')
                self.excel_path.rename(backup_path)
                self.logger.info(f"Backup criado: {backup_path}")
            
            # Salvar arquivo
            self.workbook.save(self.excel_path)
            self.logger.info(f"Excel salvo: {self.excel_path}")
            
        except Exception as e:
            self.logger.error(f"Erro ao salvar Excel: {e}")
            raise
    
    def obter_estatisticas(self) -> dict:
        """
        Obtém estatísticas do processamento
        
        Returns:
            Dicionário com estatísticas
        """
        try:
            if not self.sheet_produtos:
                return {}
            
            total = 0
            concluidos = 0
            nao_encontrados = 0
            erros = 0
            pendentes = 0
            
            # Contar produtos por status
            for row in range(2, self.sheet_produtos.max_row + 1):
                codigo = self.sheet_produtos[f'A{row}'].value
                if codigo and str(codigo).strip():
                    total += 1
                    status = self.sheet_produtos[f'U{row}'].value
                    
                    if not status or status.strip() == '':
                        pendentes += 1
                    elif status == 'CONCLUIDO':
                        concluidos += 1
                    elif status == 'NAO_ENCONTRADO':
                        nao_encontrados += 1
                    elif status == 'ERRO':
                        erros += 1
            
            return {
                'total': total,
                'concluidos': concluidos,
                'nao_encontrados': nao_encontrados,
                'erros': erros,
                'pendentes': pendentes
            }
            
        except Exception as e:
            self.logger.error(f"Erro ao obter estatísticas: {e}")
            return {}
    
    def validar_excel(self) -> List[str]:
        """
        Valida a estrutura do arquivo Excel
        
        Returns:
            Lista de problemas encontrados
        """
        problemas = []
        
        try:
            # Verificar se workbook está carregado
            if not self.workbook:
                problemas.append("Arquivo Excel não foi carregado")
                return problemas
            
            # Verificar abas obrigatórias
            if "Produtos" not in self.workbook.sheetnames:
                problemas.append("Aba 'Produtos' não encontrada")
            
            # Verificar se há produtos para processar
            if self.sheet_produtos:
                produtos_validos = 0
                for row in range(2, self.sheet_produtos.max_row + 1):
                    codigo = self.sheet_produtos[f'A{row}'].value
                    if codigo and str(codigo).strip():
                        produtos_validos += 1
                
                if produtos_validos == 0:
                    problemas.append("Nenhum código de produto encontrado na aba 'Produtos'")
            
            # Verificar credenciais se necessário
            if self.sheet_credenciais:
                username = self.sheet_credenciais['B4'].value
                password = self.sheet_credenciais['B5'].value
                
                if not username or not password:
                    problemas.append("Credenciais de login não preenchidas na aba 'Credenciais'")
            
        except Exception as e:
            problemas.append(f"Erro ao validar Excel: {e}")
        
        return problemas